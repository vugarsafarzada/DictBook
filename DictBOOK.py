from openpyxl import Workbook, load_workbook
import time
import os
import random

# CREATED BY "VUGAR SAFARZADA" (30.08.2020)

line = '------------------------------------------------------------'

lc = os.getcwd()


"""
FOR COLORS
os.chdir('C:\Windows\System32')
os.system("color f0")
os.system("mode 60")
os.chdir(lc)
"""
os.system('cls')

wb = load_workbook('data.xlsx')
ws = wb.active
#---------------------------------------------
print(f"""
{line}
                Welcome to DictBOOK dictionary!
                   **  Enjoy your world  **
{line}
* az - For to Azerbaijana language
* en - For to English language
* search - For to search words which exist on data

* add - Add new word
* del - Delete a word
* status - Take info

* data - Open data excel file

* $stop - Back to before
* $exit - Turn off DictBook
{line}
""")

#---------------------------------------------

def Az(word):
    wb = load_workbook('data.xlsx')
    ws = wb.active
    i = 1
    while (i <= ws.max_row):
        locB = "B" + str(i)
        if word == str(ws[locB].value):
            locA = "A" + str(i)
            print()
            print(ws[locA].value)
        i += 1

#---------------------------------------------

def En(word):
    wb = load_workbook('data.xlsx')
    ws = wb.active
    i = 1
    while (i <= ws.max_row):
        locA = "A" + str(i)
        if word == str(ws[locA].value):
            locB = "B" + str(i)
            print()
            print(ws[locB].value)
        i += 1
        
#---------------------------------------------

from openpyxl import Workbook, load_workbook

wb = load_workbook("data.xlsx")
ws = wb.active

#---------------------------------------------

def add(en, az, prn):

    azwords = []
    enwords = []
    pronounce = []

    num = ws.max_row

    for words in ws.iter_cols(min_row=1, min_col=1, max_row=ws.max_row, max_col=1):
        for call in words:
            enwords.append(call.value)

    for words in ws.iter_cols(min_row=1, min_col=2, max_row=ws.max_row, max_col=2):
        for call in words:
            azwords.append(call.value)

    for words in ws.iter_cols(min_row=1, min_col=3, max_row=ws.max_row, max_col=3):
        for call in words:
            pronounce.append(call.value)

    if (en in enwords) and (az in azwords):
        print(f"\nAlready there is'{az} - {enwords[azwords.index(az)]}' defined on data!")

    elif en in enwords:
        print(f"\nAlready there is '{en}' defined English word\n on data!")
        print(f"** {en} - {azwords[enwords.index(en)]} **")
        i = 1
        while i <=ws.max_row:
            locA = "A" + str(i)
            if(ws[locA].value == en):
                locB = "B" + str(i)
                while True:
                    cmd = input(f"\n - edit({ws[locA].value})/>")
                    if cmd == "$stop":
                        break
                    txt = f"{ws[locB].value}, {cmd}"
                    ws[locB] = txt
                    wb.save("data.xlsx")
                    print(f"{ws[locA].value} - {ws[locB].value}")
            i += 1

    elif az in azwords:
        num += 1
        ws.append([en, az])
        wb.save("data.xlsx")
        print("\nDone!\n")
        print(f"** {az} - {enwords[azwords.index(az)]} **")

    elif (en not in enwords) and (az not in azwords):
        num += 1
        prn = f'[{prn}]'
        ws.append([en, az, prn])
        wb.save("data.xlsx")

    print(f"\n({num}) word was completed!")
    print(f"\n{line}")

#-------------------------------------------------------------------

def delword(word):
    i = 1
    if word == '$all':
        max = ws.max_row
        quest = input("Are you sure delete all words on data? (Y/N):")
        if (quest == "Y") or (quest == "y"):
            x = 1
            y = 1
            print("Please wait...")
            while x <= ws.max_column:
                ws.delete_cols(x)
                x += 1
                time.sleep(1)
                wb.save("data.xlsx")

            while y <= ws.max_column:
                ws.delete_cols(y)
                y += 1
                time.sleep(1)
                wb.save("data.xlsx")

            print("Done!")
            print(f"Deleted {max} words!")
            print(line)
#-------------------------------------------------------------------

    while (i <= ws.max_row):
        locA = "A" + str(i)
        locB = "B" + str(i)
        if word == str(ws[locB].value):
            print(f"'{ws[locA].value}-{ws[locB].value}' deleting...")
            ws.delete_rows(i)
            wb.save("data.xlsx")
        elif word == str(ws[locA].value):
            print(f"'{ws[locA].value}-{ws[locB].value}' deleting...")
            ws.delete_rows(i)
            wb.save("data.xlsx")
        i += 1
#-------------------------------------------------------------------

def status():
    os.system('cls')
    enw = []
    azw = []
    prn = []
    for row in ws.iter_cols(min_row=1, min_col=1, max_row=ws.max_row, max_col=1):
        for cell in row:
            enw.append(cell.value)
    wb.close()
    for row in ws.iter_cols(min_row=1, min_col=2, max_row=ws.max_row, max_col=2):
        for cell in row:
            azw.append(cell.value)
    wb.close()
    for row in ws.iter_cols(min_row=1, min_col=3, max_row=ws.max_row, max_col=3):
        for cell in row:
            prn.append(cell.value)
    wb.close()
    i = 0
#-------------------------------------------------------------------

    print(f"""
There are {ws.max_row} total word on data.

* all - Look all words

* azword - Look Azerbaijana words
* enword - Look English words

* -first n - Look first n words 
* -last n - Look last n words 
* -random - Look words random
""")
#-------------------------------------------------------------------

    while True:
        cmd = input("\n- Status/>")
        if cmd == "enword":
            os.system('cls')
            print()
            print('$manual - look words manually\n')
            spd = input('Speed: ')
            print()

            if spd == "$manual":
                print(line)
                print("\nPress ENTER for next")
                print('$stop - For stop words manually\n')
                print(line)
                print()
                n = 0
                for i in enw:
                    print("*", i+prn[n])
                    n += 1
                    cm = input()
                    if cm == "$stop":
                        break
                    elif cm == "-":
                        print(f"  ({azw[n-1]})  ")
                    print(line)
            elif float(spd) >= 0:
                n = 0
                for i in enw:
                    print("*", i+prn[n])
                    n += 1
                    time.sleep(float(spd))
            else:
                print("@ERROR!")
        elif cmd == "azword":
            os.system('cls')
            print()
            print('$manual - look words manually\n')
            spd = input('Speed: ')
            print()

            
            if spd == "$manual":
                print(line)
                print("\nPress ENTER for next")
                print('$stop - For stop words manually\n')
                print(line)
                print()
                n = 0
                for i in azw:
                    n += 1
                    print("*",i)
                    cm = input()
                    if cm == "$stop":
                        break
                    elif cm == "-":
                        print(f"-- {enw[n-1]} --")
                    print(line)
            elif float(spd) >= 0:
                os.system('cls')
                for i in azw:
                    print("*",i)
                    time.sleep(float(spd))
        elif cmd == "all":
            os.system('cls')
            print()
            print('$manual - look words manually\n')
            i = 0
            n = 1
            spd = input('Speed: ')          
            print()
            os.system('cls')
            if spd == "$manual":
                print(line)
                print("\nPress ENTER for next")
                print('$stop - For stop words manually\n')
                print(line)
                print()

                while i < ws.max_row:
                    print(f"{n}. {enw[i]}{prn[i]} - {azw[i]}")
                    i += 1
                    n += 1
                    print(line)
                    cm = input()
                    if cm == "$stop":
                        break

            elif float(spd) >= 0:
                os.system('cls')
                while i < ws.max_row:
                    print(f"{n}. {enw[i]}{prn[i]} - {azw[i]}")
                    print(line)
                    time.sleep(float(spd))
                    i += 1
                    n += 1

                    
        elif cmd == "azword -random":
            os.system('cls')
            already = []
            n = 0
            print()
            while True:
                rndnum = random.randint(0,ws.max_row-1)
                
                if rndnum not in already:
                    n +=1
                    print(f"{n}. {azw[rndnum]}")
                    already.append(rndnum)
                    c=input()
                    if (n==ws.max_row) or (c == "$stop"):
                        break
                    elif c == "-":
                        print(f"** {enw[rndnum]} **")
                    print(line)

                    
        elif cmd == "enword -random":
            os.system('cls')
            already = []
            n = 0
            print()
            while True:
                rndnum = random.randint(0,ws.max_row-1)

                if rndnum not in already:
                    n +=1
                    print(f"{n}. {enw[rndnum]}{prn[rndnum]}")
                    already.append(rndnum)
                    c=input()
                    if (n==ws.max_row) or (c == "$stop"):
                        break
                    elif c == "-":
                        print(f"** {azw[rndnum]} **")    
                    print(line)

                    
        elif cmd == "all -random":
            os.system('cls')
            already = []
            n = 0
            print()
            while True:
                rndnum = random.randint(0,ws.max_row-1)

                if rndnum not in already:
                    n +=1
                    print(f"{n}. {enw[rndnum]}{prn[rndnum]} - {azw[rndnum]}")
                    print(line)
                    already.append(rndnum)
                    c = input()
                    if (n==ws.max_row) or (c == "$stop"):
                        break     

        elif cmd == "$stop":
            print(line)
            os.system('cls')
            break
#-------------------------------------------------------------------

        elif cmd.startswith("all -first "):
            try:
                num = int(cmd[11:])
                if num <= ws.max_row:
                    print('\n$manual - look words manually')
                    print('-random - Look words random\n')
                    i = 0
                    n = 1
                    spd = input('Speed: ')          
                    print()
                    os.system('cls')

                    if spd == "$manual":
                        print(line)
                        print("\nPress ENTER for next")
                        print('$stop - For stop words manually\n')
                        print(line)
                        print()
                        while i < num:
                            print(f"{n}. {enw[i]}{prn[i]} - {azw[i]}")
                            i += 1
                            n += 1
                            print(line)
                            cm = input()
                            if cm == "$stop":
                                break
                    
                    elif spd == "$manual -random":
                        already = []
                        n = 0
                        print()
                        os.system('cls')

                        while True:
                            rndnum = random.randint(0,num-1)

                            if rndnum not in already:
                                n +=1
                                print(f"{n}. {enw[rndnum]}{prn[rndnum]} - {azw[rndnum]}")
                                print(line)
                                already.append(rndnum)
                                c = input()
                                if (n==num) or (c == "$stop"):
                                    break     

                            elif cmd == "$stop":
                                print(line)
                                os.system('cls')
                                break

                    elif float(spd) >= 0:
                        os.system('cls')
                        while i < num:
                            print(f"{n}. {enw[i]}{prn[i]} - {azw[i]}")
                            print(line)
                            time.sleep(float(spd))
                            i += 1
                            n += 1

                elif num > ws.max_row:
                    print("\nThere are total 400 word on data!")      
            except:
                print(f"\n'{cmd[11:]}' isn't the number!")
#-------------------------------------------------------------------
            
        elif cmd.startswith("all -last "):
            try:
                num = int(cmd[10:])
                if num <= ws.max_row:
                    print('\n$manual - look words manually')
                    print('-random - Look words random\n')
                    i = ws.max_row-num
                    n = ws.max_row-num+1
                    spd = input('Speed: ')          
                    print()
                    os.system('cls')
                    if spd == "$manual":
                        print(line)
                        print("\nPress ENTER for next")
                        print('$stop - For stop words manually\n')
                        print(line)
                        print()
                        while i < ws.max_row:
                            print(f"{n}. {enw[i]}{prn[i]} - {azw[i]}")
                            i += 1
                            n += 1
                            print(line)
                            cm = input()
                            if cm == "$stop":
                                break
                    
                    elif spd == "$manual -random":
                        already = []
                        n = ws.max_row-num
                        print()
                        os.system('cls')

                        while True:
                            rndnum = random.randint(ws.max_row-num, ws.max_row-1)
                            if rndnum not in already:
                                n +=1
                                print(f"{n}. {enw[rndnum]}{prn[rndnum]} - {azw[rndnum]}")
                                print(line)
                                already.append(rndnum)
                                c = input()
                                if (n==ws.max_row) or (c == "$stop"):
                                    break     

                            elif cmd == "$stop":
                                print(line)
                                os.system('cls')
                                break

                    elif float(spd) >= 0:
                        os.system('cls')
                        while i < ws.max_row:
                            print(f"{n}. {enw[i]}{prn[i]} - {azw[i]}")
                            print(line)
                            time.sleep(float(spd))
                            i += 1
                            n += 1

                elif num > ws.max_row:
                    print("\nThere are total 400 word on data!")      
            except:
                print("@ERROR")
#-------------------------------------------------------------------

        elif cmd.startswith("enword -first "):
            try:
                num = int(cmd[14:])
                if num <= ws.max_row:
                    print('\n$manual - look words manually')
                    print('-random - Look words random\n')
           
                    spd = input('Speed: ')
                    print()

                    if spd == "$manual":
                        os.system('cls')
                        print("Press ENTER for next")
                        print('$stop - For stop words manually\n')
                        print(line)
                        n = 0
                        for i in enw:
                            n+=1 
                            print(f"{n}.",i+prn[n-1])
                            cm = input()
                            if cm == "-":
                                print(f"** {azw[n-1]} **")
                            if (cm == "$stop") or (n == num):
                                break
                      
                            print(line)
                    
                    elif spd == "$manual -random":
                        os.system('cls')
                        print("Press ENTER for next")
                        print('$stop - For stop words manually\n')
                        print(line)
                        already = []
                        n = 0
                        print()
                        while True:
                            rndnum = random.randint(0,num-1)

                            if rndnum not in already:
                                n +=1
                                print(f"{n}. {enw[rndnum]}{prn[rndnum]}")
                                already.append(rndnum)
                                c=input()
                                if c == "-":
                                    print(f"** {azw[rndnum]} **") 
                                if (n==num) or (c == "$stop"):
                                    break
                                   
                                print(line)
                    elif float(spd) >= 0:
                        os.system('cls')
                        n = 0
                        for i in enw:
                            n+=1
                            print(f"{n}.", i+prn[n-1])
                            print(line)
                            time.sleep(float(spd))
                            if n == num:
                                break
            except:
                print("@ERROR")
#-------------------------------------------------------------------

        elif cmd.startswith("azword -first "):
            try:
                num = int(cmd[14:])
                if num <= ws.max_row:
                    print('\n$manual - look words manually')
                    print('-random - Look words random\n')
           
                    spd = input('Speed: ')
                    print()
                    if spd == "$manual":
                        os.system('cls')
                        print("Press ENTER for next")
                        print('$stop - For stop words manually\n')
                        print(line)
                        n = 0
                        for i in azw:
                            n+=1 
                            print(f"{n}.",i)
                            cm = input()
                            if cm == "-":
                                print(f"** {enw[n-1]} **")
                            if (cm == "$stop") or (n == num):
                                break
                      
                            print(line)
                    
                    elif spd == "$manual -random":
                        os.system('cls')
                        print("Press ENTER for next")
                        print('$stop - For stop words manually\n')
                        print(line)
                        already = []
                        n = 0
                        print()
                        while True:
                            rndnum = random.randint(0,num-1)

                            if rndnum not in already:
                                n +=1
                                print(f"{n}. {azw[rndnum]}")
                                already.append(rndnum)
                                c=input()
                                if c == "-":
                                    print(f"** {enw[rndnum]} **") 
                                if (n==num) or (c == "$stop"):
                                    break
                                print(line)

                    elif float(spd) >= 0:
                        os.system('cls')
                        n = 0
                        for i in azw:
                            n+=1
                            print(f"{n}.",i)
                            print(line)
                            time.sleep(float(spd))
                            if n == num:
                                break
            except:
                print("@ERROR")

#-------------------------------------------------------------------

        elif cmd.startswith("azword -last "):
            try:
                num = int(cmd[13:])
                if num <= ws.max_row:
                    print('\n$manual - look words manually')
                    print('-random - Look words random\n')
           
                    spd = input('Speed: ')
                    print()

                    if spd == "$manual":
                        os.system('cls')
                        print("Press ENTER for next")
                        print('$stop - For stop words manually\n')
                        print(line)
                        i = ws.max_row-num-1
                        n = ws.max_row-num

                        while i < ws.max_row:
                            n+=1
                            i+=1
                            print(f"{n}.",azw[i])
                            c = input()
                            if c == "-":
                                print(f"** {enw[i]} **")
                            if (n == ws.max_row) or (c == "$stop"):
                                break
                            print(line)
                    
                    elif spd == "$manual -random":
                        os.system('cls')
                        print("Press ENTER for next")
                        print('$stop - For stop words manually\n')
                        print(line)
                        already = []
                        n = ws.max_row-num
                        print()

                        while True:
                            rndnum = random.randint(ws.max_row-num, ws.max_row-1)

                            if rndnum not in already:
                                n +=1
                                print(f"{n}. {azw[rndnum]}")
                                already.append(rndnum)
                                c=input()
                                if c == "-":
                                    print(f"** {enw[rndnum]} **") 
                                if (n==ws.max_row) or (c == "$stop"):
                                    break
                                print(line)

                    elif float(spd) >= 0:
                        os.system('cls')
                        i = ws.max_row-num-1
                        n = ws.max_row-num
                        while i < ws.max_row:
                            n+=1
                            i+=1
                            print(f"{n}.",azw[i])
                            print(line)
                            time.sleep(float(spd))
                            if n == ws.max_row:
                                break
            except:
                print("@ERROR")
#-------------------------------------------------------------------

        elif cmd.startswith("enword -last "):
            try:
                num = int(cmd[13:])
                if num <= ws.max_row:
                    print('\n$manual - look words manually')
                    print('-random - Look words random\n')
           
                    spd = input('Speed: ')
                    print()

                    if spd == "$manual":
                        os.system('cls')
                        print("Press ENTER for next")
                        print('$stop - For stop words manually\n')
                        print(line)
                        i = ws.max_row-num-1
                        n = ws.max_row-num
                        while i < ws.max_row:
                            n+=1
                            i+=1
                            print(f"{n}.",enw[i]+prn[i])
                            c = input()
                            if c == "-":
                                print(f"** {azw[i]} **")
                            if (n == ws.max_row) or (c == "$stop"):
                                break
                            print(line)
                    
                    elif spd == "$manual -random":
                        os.system('cls')
                        print("Press ENTER for next")
                        print('$stop - For stop words manually\n')
                        print(line)
                        already = []
                        n = ws.max_row-num
                        print()
                        while True:
                            rndnum = random.randint(ws.max_row-num, ws.max_row-1)

                            if rndnum not in already:
                                n +=1
                                print(f"{n}. {enw[rndnum]}{prn[rndnum]}")
                                already.append(rndnum)
                                c=input()
                                if c == "-":
                                    print(f"** {azw[rndnum]} **") 
                                if (n==ws.max_row) or (c == "$stop"):
                                    break
                                print(line)

                    elif float(spd) >= 0:
                        os.system('cls')
                        i = ws.max_row-num-1
                        n = ws.max_row-num
                        while i < ws.max_row:
                            n+=1
                            i+=1
                            print(f"{n}.",enw[i] + prn[i])
                            print(line)
                            time.sleep(float(spd))
                            if n == ws.max_row:
                                break
            except:
                print("@ERROR")

#-------------------------------------------------------------------

def search():

    azwords = []
    enwords = []
    pronounce = []

    for words in ws.iter_cols(min_row=1, min_col=1, max_row=ws.max_row, max_col=1):
        for call in words:
            enwords.append(call.value)

    for words in ws.iter_cols(min_row=1, min_col=2, max_row=ws.max_row, max_col=2):
        for call in words:
            azwords.append(call.value)
    
    for words in ws.iter_cols(min_row=1, min_col=3, max_row=ws.max_row, max_col=3):
        for call in words:
            pronounce.append(call.value)
    
    print("* $stop - For to stop searching")
    print("* ... $en - For search on English words only")
    print("* ... $az - For search on Azerbaijani words only")

    while True:

        key = input("\nSearch: ")

        os.system('cls')

        if key == "$stop":
            break

        print(f"Searching '{key}' ...")
        time.sleep(1)
        findsAz = []
        findsEn = []

        for azw in azwords:
            if azw.startswith(key):
                findsAz.append(azw)

        for enw in enwords:
            if enw.startswith(key):
                findsEn.append(enw)

        if key.endswith(' $en'):
            for enw in enwords:
                if enw.startswith(key[:len(key)-4]):
                    findsEn.append(enw)
            print("\n----------EN----------\n")    
            if len(findsEn) == 0:
                print("** Not found **")
            else:
                for i in findsEn:
                    nEn = enwords.index(i)
                    print(f"{i}{pronounce[nEn]} - {azwords[nEn]}")
                
        elif key.endswith(' $az'):
            for azw in azwords:
                if azw.startswith(key[:len(key)-4]):   
                     findsAz.append(azw)           
            print("\n----------AZ----------\n")    
            if len(findsAz) == 0:
                print("** Not found **")
            else:
                for i in findsAz:
                    nAz = azwords.index(i)
                    print(f"{i} - {enwords[nAz]}{pronounce[nAz]}")
                

        else:
            print("\n------------EN------------\n")    
            if len(findsEn) == 0:
                print("** Not found **")
            else:
                for i in findsEn:
                    nEn = enwords.index(i)
                    print(f"{i}{pronounce[nEn]} - {azwords[nEn]}")
            
            print("\n------------AZ------------\n")
            if len(findsAz) == 0:
                print("** Not found **")
            else:
                for i in findsAz:
                    nAz = azwords.index(i)
                    print(f"{i} - {enwords[nAz]}{pronounce[nAz]}")

        print("\n---------------------------")

#-------------------------------------------------------------------

while True:
    
    cmd = input("\nDICTBOOK/>")

    wb = load_workbook('data.xlsx')
    ws = wb.active

    if cmd == "az":
        os.system('cls')
        while True:
            cmdAz = input("\n- AZ/>")
            if cmdAz == "$stop":
                print(line)
                break
            Az(cmdAz)

#---------------------------------------------

    elif cmd == "en":
        os.system('cls')
        while True:
            cmdEn = input("\n- EN/>")
            if cmdEn == "$stop":
                print(line)
                break
            En(cmdEn)

#---------------------------------------------

    elif cmd == "add":
        os.system('cls')
        
        while True:
            newEn = input("\n- ADD (en)/>")
            if newEn == "$stop":
                print(line)
                break
            newPrn = input(f"\n- ADD (pronounce)/>")
            if newPrn == "$stop":
                print(line)
                break
            newAz = input("\n- ADD (az)/>")
            if newAz == "$stop":
                print(line)
                break
           
            add(newEn, newAz, newPrn)
            
#---------------------------------------------

    elif cmd == "del":
        print("\n$all - For to delete all file")
        while True:
            dword = input("\n- DEL/>")
            if dword == "$stop":
                print(line)
                break
            delword(dword)

#---------------------------------------------

    elif cmd == "status":
        os.system('cls')
        status()

#---------------------------------------------

    elif cmd == "search":
        os.system('cls')
        search()

#---------------------------------------------

    elif cmd == "data":
        try:
            os.startfile('data.xlsx')
            print("Running...")
        except:
            print("data.xlsx can't find!")

#---------------------------------------------

    elif cmd == "$exit":
        print(line)
        break